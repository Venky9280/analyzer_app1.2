# Your imports and setup remain the same
import streamlit as st
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import statsmodels.api as sm
import statsmodels.formula.api as smf
import plotly.express as px
from plotly.graph_objs import Figure
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans
import numpy as np
import io
import tempfile
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

st.set_page_config(page_title="Super Data Analyst Bot", layout="wide")
st.title("📊 Super Data Analyst Bot")
st.markdown("""
<style>
    .main { background-color: #f7f9fc; }
    .stSelectbox, .stButton { font-size: 16px; }
    .insight-box { padding: 15px; background: #f0f2f6; border-radius: 10px; margin: 10px 0; }
</style>
""", unsafe_allow_html=True)

def load_excel_file(uploaded_file):
    xls = pd.ExcelFile(uploaded_file)
    sheet_names = xls.sheet_names
    selected_sheet = st.selectbox("📁 Select a sheet to analyze", sheet_names)
    df_raw = xls.parse(selected_sheet, header=None)
    header_row_index = df_raw.notnull().sum(axis=1).idxmax()
    return xls.parse(selected_sheet, header=header_row_index)

def generate_excel_report(df, figures, analysis_results, corr_df=None):
    output = io.BytesIO()
    wb = Workbook()
    
    wb.create_sheet("Raw Data")
    wb["Raw Data"].append(list(df.columns))
    for row in df.itertuples(index=False):
        wb["Raw Data"].append(row)
        
    wb.create_sheet("Descriptive Stats")
    stats_df = df.describe(include='all').reset_index()
    wb["Descriptive Stats"].append(list(stats_df.columns))
    for row in stats_df.itertuples(index=False):
        wb["Descriptive Stats"].append(row)
    
    if corr_df is not None:
        wb.create_sheet("Correlation Results")
        wb["Correlation Results"].append(list(corr_df.columns))
        for row in corr_df.itertuples(index=False):
            wb["Correlation Results"].append(row)
    
    with tempfile.TemporaryDirectory() as tmpdir:
        img_paths = []
        for idx, (fig, title) in enumerate(figures):
            path = Path(tmpdir) / f"figure_{idx}.png"
            if isinstance(fig, plt.Figure):
                fig.savefig(path, bbox_inches='tight', dpi=150)
            elif isinstance(fig, Figure):
                fig.write_image(path, scale=2)
            img_paths.append((path, title))
        
        ws_images = wb.create_sheet("Graphs")
        row_idx = 1
        for path, title in img_paths:
            img = XLImage(path)
            ws_images.add_image(img, f'A{row_idx}')
            ws_images[f'B{row_idx}'] = title
            row_idx += img.height // 20 + 2
    
    ws_analysis = wb.create_sheet("Analysis Insights")
    ws_analysis.append(["Analysis Type", "Insights"])
    for result in analysis_results:
        ws_analysis.append([result['type'], result['text']])
    
    wb.save(output)
    return output.getvalue()

uploaded_file = st.file_uploader("📂 Upload a CSV or Excel file", type=["csv", "xlsx", "xls"])

if uploaded_file:
    if 'report_figures' not in st.session_state:
        st.session_state.report_figures = []
    if 'analysis_results' not in st.session_state:
        st.session_state.analysis_results = []
    if 'corr_df' not in st.session_state:
        st.session_state.corr_df = None
    
    file_ext = uploaded_file.name.split('.')[-1].lower()
    df = pd.read_csv(uploaded_file) if file_ext == "csv" else load_excel_file(uploaded_file)

    if df is not None and not df.empty:
        df.columns = df.columns.astype(str)

        st.subheader("🔍 Data Preview")
        st.dataframe(df, use_container_width=True)
        st.markdown(f"**Shape:** {df.shape[0]} rows × {df.shape[1]} columns")

        st.subheader("🧹 Data Cleaning")
        with st.expander("Click to clean data"):
            st.markdown("""
            **Cleaning Strategy Guide:**
            - **Drop NA**: Safe if <5% missing randomly
            - **Fill with 0**: Appropriate for true zeros
            - **Fill Missing**: Better to use mean/median for normal distributions
            """)
            if st.checkbox("Drop rows with missing values"):
                df = df.dropna()
            if st.checkbox("Fill missing numeric values with 0"):
                df = df.fillna({col: 0 for col in df.select_dtypes(include="number").columns})
            if st.checkbox("Replace blanks in categorical data with 'Missing'"):
                df = df.fillna("Missing").replace("", "Missing")

        st.subheader("📊 Descriptive Statistics")
        st.dataframe(df.describe(include='all'), use_container_width=True)

        st.subheader("📈 Correlation Heatmap & Interpretation")
        numeric_df_corr = df.select_dtypes(include='number')
        if not numeric_df_corr.empty:
            corr = numeric_df_corr.corr()
            fig1, ax1 = plt.subplots(figsize=(10, 6))
            sns.heatmap(corr, annot=True, cmap="coolwarm", ax=ax1)
            st.pyplot(fig1)
            st.session_state.report_figures.append((fig1, "Correlation Heatmap"))

            strong_corrs = []
            for i, col1 in enumerate(corr.columns):
                for j, col2 in enumerate(corr.columns):
                    if i < j and abs(corr.loc[col1, col2]) > 0.7:
                        strong_corrs.append({
                            "Feature 1": col1,
                            "Feature 2": col2,
                            "Correlation": corr.loc[col1, col2],
                            "Use Case": f"{col2} might be predicted using {col1}" if corr.loc[col1, col2] > 0 
                                        else f"{col1} and {col2} are inversely related"
                        })

            if strong_corrs:
                st.session_state.corr_df = pd.DataFrame(strong_corrs).sort_values(
                    by="Correlation", 
                    key=abs, 
                    ascending=False
                ).head(10)

                st.markdown("### 🔍 Top Correlations (|r| > 0.7)")
                st.dataframe(
                    st.session_state.corr_df.style.format({"Correlation": "{:.4f}"}),
                    use_container_width=True,
                    height=400
                )
                st.download_button(
                    "📥 Download Correlation Insights",
                    st.session_state.corr_df.to_csv(index=False),
                    file_name="top_correlations.csv"
                )
            else:
                st.info("No strong correlations found (|r| > 0.7)")
        else:
            st.info("No numeric columns for correlation analysis")

        st.subheader("📊 Auto Graph Generator")
        col1, col2, col3 = st.columns(3)
        with col1:
            x_var = st.selectbox("Select X variable", df.columns)
        with col2:
            y_var = st.selectbox("Select Y variable", [None] + df.columns.tolist())
        with col3:
            sort_order = st.selectbox("Sort Order", ["None", "Ascending", "Descending"])

        chart_type = st.selectbox("Chart type", ["Bar Chart", "Histogram", "Scatterplot", "Boxplot", "Line Chart", "Pie Chart"])
        try:
            if chart_type == "Bar Chart":
                bar_data = df[x_var].value_counts().reset_index()
                bar_data.columns = [x_var, 'count']
                if sort_order == "Ascending":
                    bar_data = bar_data.sort_values(by='count')
                elif sort_order == "Descending":
                    bar_data = bar_data.sort_values(by='count', ascending=False)
                fig = px.bar(bar_data, x=x_var, y='count', title=f"Bar Chart of {x_var}")
            elif chart_type == "Histogram":
                fig = px.histogram(df, x=x_var, title=f"Histogram of {x_var}")
            elif chart_type == "Scatterplot" and y_var:
                fig = px.scatter(df, x=x_var, y=y_var, title=f"Scatterplot: {x_var} vs {y_var}")
            elif chart_type == "Boxplot" and y_var:
                fig = px.box(df, x=y_var, y=x_var, title=f"Boxplot of {x_var} vs {y_var}")
            elif chart_type == "Line Chart" and y_var:
                fig = px.line(df, x=x_var, y=y_var, title=f"Line Chart: {x_var} vs {y_var}")
            elif chart_type == "Pie Chart":
                pie_data = df[x_var].value_counts()
                fig = px.pie(pie_data, names=pie_data.index, values=pie_data.values, title=f"Pie Chart of {x_var}")
            st.plotly_chart(fig)
            st.session_state.report_figures.append((fig, f"Graph - {chart_type}"))
        except Exception as e:
            st.error(f"Error generating chart: {e}")

        st.subheader("🧠 Clustering")
        numeric_df_clean = df.select_dtypes(include='number').dropna()
        scaled_data = StandardScaler().fit_transform(numeric_df_clean)

        if scaled_data.shape[0] < 3:
            st.warning("Not enough rows for clustering. Need at least 3.")
        else:
            user_k = st.slider("Select number of clusters (K)", 2, 10, 3)
            kmeans = KMeans(n_clusters=user_k, random_state=0).fit(scaled_data)

            # ✅ Robust fix for index mismatch
            cluster_series = pd.Series(kmeans.labels_, index=numeric_df_clean.index, name='Cluster')
            df = df.join(cluster_series, how='left')

            st.subheader("🔍 Cluster Insights")
            st.dataframe(df[['Cluster']].head())

        st.subheader("📥 Download Analysis Report")
        report_bytes = generate_excel_report(
            df, 
            st.session_state.report_figures, 
            st.session_state.analysis_results,
            st.session_state.corr_df
        )
        st.download_button(
            label="Download Excel Report",
            data=report_bytes,
            file_name="data_analysis_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
